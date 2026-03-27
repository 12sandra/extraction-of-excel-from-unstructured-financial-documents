import logging
import os
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger("gst2_fastapi.excel_generator")

# GST 2A Excel column headers (matches official GST 2A format)
GST2A_COLUMNS = [
    "GSTIN of Supplier",
    "Trade/Legal Name",
    "Invoice Number",
    "Invoice Date",
    "Invoice Value (₹)",
    "Place of Supply",
    "Reverse Charge",
    "Invoice Type",
    "Rate (%)",
    "Taxable Value (₹)",
    "IGST (₹)",
    "CGST (₹)",
    "SGST/UTGST (₹)",
    "Cess (₹)",
    "Return Period",
    "Source",
    "Confidence Score"
]

# Field mapping from our extracted fields to column names
FIELD_TO_COLUMN = {
    "GSTIN": "GSTIN of Supplier",
    "TRADE_NAME": "Trade/Legal Name",
    "INVOICE_NO": "Invoice Number",
    "INVOICE_DATE": "Invoice Date",
    "INVOICE_VALUE": "Invoice Value (₹)",
    "PLACE_OF_SUPPLY": "Place of Supply",
    "TAXABLE_VALUE": "Taxable Value (₹)",
    "IGST": "IGST (₹)",
    "CGST": "CGST (₹)",
    "SGST": "SGST/UTGST (₹)",
    "CESS": "Cess (₹)",
    "RETURN_PERIOD": "Return Period",
}


class ExcelGenerator:
    """Creates formatted Excel files from extracted GST 2A data."""
    
    # Color scheme
    HEADER_COLOR = "1F4E79"    # Dark blue
    SUBHEADER_COLOR = "2E75B6" # Medium blue
    ROW_ALT_COLOR = "DEEAF1"   # Light blue alternate rows
    WHITE = "FFFFFF"
    GREEN = "70AD47"           # For verified data
    YELLOW = "FFC000"          # For uncertain data
    RED = "FF0000"             # For errors
    
    def generate(
        self,
        extracted_data: List[Dict[str, Any]],
        output_path: str,
        file_name: str,
        job_id: str
    ) -> str:
        """
        Generate a formatted Excel file from extracted GST data.
        
        Args:
            extracted_data: List of invoice records (one dict per invoice row)
            output_path: Directory to save Excel file
            file_name: Original uploaded file name
            job_id: Unique job ID
            
        Returns:
            Path to the generated Excel file
        """
        output_path = Path(output_path)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"GST2A_{Path(file_name).stem}_{timestamp}.xlsx"
        excel_path = output_path / excel_filename
        
        wb = Workbook()
        
        # Create main data sheet
        self._create_data_sheet(wb, extracted_data, file_name)
        
        # Create summary sheet
        self._create_summary_sheet(wb, extracted_data, file_name)
        
        # Create metadata sheet
        self._create_metadata_sheet(wb, file_name, job_id, extracted_data)
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        wb.save(str(excel_path))
        logger.info(f"Excel saved: {excel_path}")
        return str(excel_path)
    
    def _create_data_sheet(self, wb: Workbook, data: List[Dict], file_name: str):
        """Create main GST 2A data sheet."""
        ws = wb.create_sheet("GST 2A Data", 0)
        
        # Title row
        ws.merge_cells("A1:Q1")
        title_cell = ws["A1"]
        title_cell.value = f"GST Form 2A — Extracted from: {file_name}"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=self.WHITE)
        title_cell.fill = PatternFill("solid", fgColor=self.HEADER_COLOR)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Subtitle
        ws.merge_cells("A2:Q2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color=self.WHITE)
        subtitle_cell.fill = PatternFill("solid", fgColor=self.SUBHEADER_COLOR)
        subtitle_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18
        
        # Column headers
        header_row = 3
        for col_idx, col_name in enumerate(GST2A_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(name="Calibri", size=10, bold=True, color=self.WHITE)
            cell.fill = PatternFill("solid", fgColor=self.HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()
        ws.row_dimensions[header_row].height = 35
        
        # Data rows
        df = self._extracted_data_to_dataframe(data)
        
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=header_row + 1):
            is_alternate = (row_idx - header_row) % 2 == 0
            row_color = self.ROW_ALT_COLOR if is_alternate else self.WHITE
            
            for col_idx, (col_name, value) in enumerate(zip(GST2A_COLUMNS, row_data), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.border = self._thin_border()
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Highlight confidence column
                if col_name == "Confidence Score" and value:
                    try:
                        conf = float(str(value).replace("%", ""))
                        if conf >= 90:
                            cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green
                        elif conf >= 75:
                            cell.fill = PatternFill("solid", fgColor="FFEB9C")  # Yellow
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
                    except:
                        pass
        
        # Auto-fit column widths
        column_widths = [22, 25, 18, 14, 16, 16, 12, 14, 8, 16, 12, 12, 14, 10, 12, 10, 14]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header rows
        ws.freeze_panes = "A4"
        
        # Auto-filter
        ws.auto_filter.ref = f"A3:Q{header_row + len(df)}"
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict], file_name: str):
        """Create a summary/totals sheet."""
        ws = wb.create_sheet("Summary")
        
        ws["A1"] = "GST 2A Summary"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=self.WHITE)
        ws["A1"].fill = PatternFill("solid", fgColor=self.HEADER_COLOR)
        ws.merge_cells("A1:C1")
        
        df = self._extracted_data_to_dataframe(data)
        
        summary_data = [
            ["Metric", "Value", "Notes"],
            ["Total Invoices", len(df), "Number of invoices extracted"],
            ["Total Invoice Value (₹)", self._safe_sum(df, "Invoice Value (₹)"), "Sum of all invoice values"],
            ["Total Taxable Value (₹)", self._safe_sum(df, "Taxable Value (₹)"), ""],
            ["Total IGST (₹)", self._safe_sum(df, "IGST (₹)"), ""],
            ["Total CGST (₹)", self._safe_sum(df, "CGST (₹)"), ""],
            ["Total SGST/UTGST (₹)", self._safe_sum(df, "SGST/UTGST (₹)"), ""],
            ["Total Cess (₹)", self._safe_sum(df, "Cess (₹)"), ""],
            ["Average Confidence", f"{df['Confidence Score'].mean():.1f}%" if 'Confidence Score' in df.columns else "N/A", "OCR accuracy"],
        ]
        
        for row_idx, row in enumerate(summary_data, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 2:
                    cell.font = Font(bold=True, color=self.WHITE)
                    cell.fill = PatternFill("solid", fgColor=self.SUBHEADER_COLOR)
                cell.border = self._thin_border()
                cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
    
    def _create_metadata_sheet(self, wb: Workbook, file_name: str, job_id: str, data: List[Dict]):
        """Create processing metadata sheet."""
        ws = wb.create_sheet("Processing Info")
        
        meta = [
            ["Field", "Value"],
            ["Source File", file_name],
            ["Job ID", job_id],
            ["Processing Date", datetime.now().strftime("%d/%m/%Y")],
            ["Processing Time", datetime.now().strftime("%H:%M:%S")],
            ["Records Extracted", len(data)],
            ["Tool Used", "GST2A Converter v1.0"],
            ["OCR Engine", "PaddleOCR + GOT-OCR 2.0"],
            ["Structure Model", "LayoutLMv3"],
            ["Validator", "LLaMA 3 8B"],
        ]
        
        for row_idx, row in enumerate(meta, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, color=self.WHITE)
                    cell.fill = PatternFill("solid", fgColor=self.HEADER_COLOR)
                cell.border = self._thin_border()
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
    
    def _extracted_data_to_dataframe(self, data: List[Dict]) -> pd.DataFrame:
        """Convert extracted field dicts to a DataFrame matching GST2A columns."""
        rows = []
        
        for record in data:
            row = {}
            for field_key, col_name in FIELD_TO_COLUMN.items():
                values = record.get(field_key, [])
                if isinstance(values, list):
                    row[col_name] = values[0] if values else ""
                else:
                    row[col_name] = values or ""
            
            # Fill remaining columns
            row["Reverse Charge"] = record.get("REVERSE_CHARGE", "N")
            row["Invoice Type"] = record.get("INVOICE_TYPE", "Regular")
            row["Rate (%)"] = record.get("TAX_RATE", "")
            row["Source"] = record.get("_source", "OCR")
            row["Confidence Score"] = f"{record.get('_confidence', 0) * 100:.1f}%"
            
            rows.append(row)
        
        if not rows:
            # Empty dataframe with correct columns
            return pd.DataFrame(columns=GST2A_COLUMNS)
        
        df = pd.DataFrame(rows)
        
        # Ensure all required columns exist
        for col in GST2A_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        
        return df[GST2A_COLUMNS]
    
    def _safe_sum(self, df: pd.DataFrame, col: str) -> str:
        """Safely sum a currency column."""
        if col not in df.columns:
            return "0.00"
        try:
            values = df[col].apply(
                lambda x: float(str(x).replace(",", "").replace("₹", "").strip()) if x else 0
            )
            return f"{values.sum():,.2f}"
        except:
            return "0.00"
    
    def _thin_border(self) -> Border:
        """Create thin cell border."""
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)


# Singleton
_excel_generator_instance = None

def get_excel_generator() -> ExcelGenerator:
    global _excel_generator_instance
    if _excel_generator_instance is None:
        _excel_generator_instance = ExcelGenerator()
    return _excel_generator_instance